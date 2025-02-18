import re
import json
import time
import pandas as pd
import requests
from bs4 import BeautifulSoup
from pathlib import Path
from tqdm import tqdm
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

class AbilityProcessor:
    def __init__(self):
        self.ability_cache_file = Path('json/ability_translations.json')
        self.ability_cache = self._load_cache()
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })

    def _load_cache(self) -> dict:
        """加载特性翻译缓存"""
        if self.ability_cache_file.exists():
            return json.loads(self.ability_cache_file.read_text(encoding='utf-8'))
        return {}

    def _save_cache(self):
        """保存特性翻译缓存"""
        self.ability_cache_file.write_text(
            json.dumps(self.ability_cache, ensure_ascii=False, indent=2),
            encoding='utf-8'
        )

    def _get_ability_translation(self, ability_en: str) -> str:
        """获取特性的中文翻译"""
        if ability_en in self.ability_cache:
            return self.ability_cache[ability_en]

        try:
            url = f"https://wiki.52poke.com/wiki/{ability_en}"
            response = self.session.get(url, timeout=10)
            soup = BeautifulSoup(response.text, 'html.parser')
            title = soup.find('h1', {'id': 'firstHeading'})
            if title:
                ability_cn = title.text.split('（')[0].strip()
                self.ability_cache[ability_en] = ability_cn
                self._save_cache()
                return ability_cn
            time.sleep(1)
        except Exception as e:
            print(f"获取特性翻译失败 {ability_en}: {e}")
        return ability_en

    def parse_ability_data(self, content: str) -> List[Dict]:
        """解析特性数据"""
        abilities = []
        
        # 匹配 flags 后面的简单数据部分
        pattern = r'flags:\s*{[^}]*},\s*name:\s*"([^"]+)",\s*rating:\s*([\d.]+),\s*num:\s*([-\d]+)'
        matches = list(re.finditer(pattern, content))

        print("解析特性数据...")
        for match in tqdm(matches, desc="处理特性"):
            try:
                name = match.group(1)
                rating = float(match.group(2))
                num = int(match.group(3))
                
                ability_data = {
                    'name_en': name,
                    'name_cn': self._get_ability_translation(name),
                    'rating': rating,
                    'num': num
                }
                abilities.append(ability_data)

            except Exception as e:
                print(f"解析错误: {e}")
                continue

        return abilities

    def process_data(self, input_file: str, output_file: str):
        """处理数据主函数"""
        print("读取数据文件...")
        with open(input_file, 'r', encoding='utf-8') as file:
            content = file.read()

        # 解析数据
        abilities = self.parse_ability_data(content)
        df = pd.DataFrame(abilities)

        # 重命名列
        column_rename = {
            'name_cn': '特性名称（中文）',
            'name_en': '特性名称（英文）',
            'rating': '评分',
            'num': '编号'
        }

        # 选择并重命名列
        df = df.rename(columns=column_rename)

        # 确保输出目录存在
        Path(output_file).parent.mkdir(exist_ok=True)

        # 使用 openpyxl 保存并设置列宽
        wb = Workbook()
        ws = wb.active

        # 写入表头
        headers = list(column_rename.values())
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # 写入数据
        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # 设置列宽
        column_widths = {
            '特性名称（中文）': 20,
            '特性名称（英文）': 25,
            '评分': 8,
            '编号': 8
        }

        for col, header in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col)].width = column_widths[header]

        # 保存文件
        print(f"保存数据到 {output_file}")
        wb.save(output_file)

def main():
    processor = AbilityProcessor()
    processor.process_data('data/abilities.ts', 'output/ability_data.xlsx')

if __name__ == "__main__":
    main() 