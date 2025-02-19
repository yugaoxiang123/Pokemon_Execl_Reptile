from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def create_ability_index_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Index"
    
    # 添加表头
    headers = ["模式", "表类型", "表文件名"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    # 添加示例数据
    data = [
        ["类型表", "", "ability_Type.xlsx"],
        ["数据表", "ability_data", "ability_data.xlsx"]
    ]
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save("C:/Users/90953/Downloads/my-pokemon-excel-designer/ability_Index.xlsx")

def create_ability_type_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Type"
    
    # 添加表头
    headers = ["种类", "对象类型", "标识名", "字段名", "字段类型", "数组切割", "值", "索引", "标记"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    # 添加示例数据
    data = [
        ["表头", "ability_data", "特性名称（英文）", "Name", "string", "", "", "", ""],
        ["表头", "ability_data", "特性名称（中文）", "ChineseName", "string", "", "", "", ""],
        ["表头", "ability_data", "世代", "Gen", "int", "", "", "", ""],
        ["表头", "ability_data", "评分", "Score", "float", "", "", "", ""],
        ["表头", "ability_data", "编号", "ID", "int", "", "", "", ""],
        ["表头", "ability_data", "特性描述", "Description", "string", "", "", "", ""]
    ]
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save("C:\\Users\\90953\\Downloads\\my-pokemon-excel-designer\\ability_Type.xlsx")

def create_item_index_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Index"
    
    # 添加表头
    headers = ["模式", "表类型", "表文件名"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    # 添加数据
    data = [
        ["类型表", "", "item_Type.xlsx"],
        ["数据表", "item_data", "item_data.xlsx"]
    ]
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save("C:/Users/90953/Downloads/my-pokemon-excel-designer/item_Index.xlsx")

def create_item_type_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Type"
    
    # 添加表头
    headers = ["种类", "对象类型", "标识名", "字段名", "字段类型", "数组切割", "值", "索引", "标记"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    # 添加数据
    data = [
        ["表头", "item_data", "道具名称（英文）", "Name", "string", "", "", "", ""],
        ["表头", "item_data", "道具名称（中文）", "ChineseName", "string", "", "", "", ""],
        ["表头", "item_data", "世代", "Gen", "int", "", "", "", ""],
        ["表头", "item_data", "编号", "ID", "int", "", "", "", ""],
        ["表头", "item_data", "精灵图编号", "SpriteID", "int", "", "", "", ""],
        ["表头", "item_data", "道具描述", "Description", "string", "", "", "", ""]
    ]
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save("C:\\Users\\90953\\Downloads\\my-pokemon-excel-designer\\item_Type.xlsx")

def create_move_index_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Index"
    
    # 添加表头
    headers = ["模式", "表类型", "表文件名"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    # 添加数据
    data = [
        ["类型表", "", "move_Type.xlsx"],
        ["数据表", "move_data", "move_data.xlsx"]
    ]
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save("C:/Users/90953/Downloads/my-pokemon-excel-designer/move_Index.xlsx")

def create_move_type_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Type"
    
    # 添加表头
    headers = ["种类", "对象类型", "标识名", "字段名", "字段类型", "数组切割", "值", "索引", "标记"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    # 添加数据
    data = [
        ["表头", "move_data", "技能名称（英文）", "Name", "string", "", "", "", ""],
        ["表头", "move_data", "技能名称（中文）", "ChineseName", "string", "", "", "", ""],
        ["表头", "move_data", "世代", "Gen", "int", "", "", "", ""],
        ["表头", "move_data", "编号", "ID", "int", "", "", "", ""],
        ["表头", "move_data", "命中率", "Accuracy", "int", "", "", "", ""],
        ["表头", "move_data", "威力", "Power", "int", "", "", "", ""],
        ["表头", "move_data", "PP值", "PP", "int", "", "", "", ""],
        ["表头", "move_data", "优先度", "Priority", "int", "", "", "", ""],
        ["表头", "move_data", "技能目标", "MoveTarget", "string", "", "", "", ""],
        ["表头", "move_data", "分类", "Classify", "string", "", "", "", ""],
        ["表头", "move_data", "技能描述", "Description", "string", "", "", "", ""],
    ]
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save("C:\\Users\\90953\\Downloads\\my-pokemon-excel-designer\\move_Type.xlsx")

def create_pokemon_index_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Index"
    
    # 添加表头
    headers = ["模式", "表类型", "表文件名"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    # 添加数据
    data = [
        ["类型表", "", "pokemon_Type.xlsx"],
        ["数据表", "pokemon_data", "pokemon_data.xlsx"]
    ]
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save("C:/Users/90953/Downloads/my-pokemon-excel-designer/pokemon_Index.xlsx")

def create_pokemon_type_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Type"
    
    # 添加表头
    headers = ["种类", "对象类型", "标识名", "字段名", "字段类型", "数组切割", "值", "索引", "标记"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    # 添加数据
    data = [
        ["表头", "pokemon_data", "编号", "ID", "int", "", "", "", ""],
        ["表头", "pokemon_data", "英文名称", "Name", "string", "", "", "", ""],
        ["表头", "pokemon_data", "中文名称", "ChineseName", "string", "", "", "", ""],
        ["表头", "pokemon_data", "世代", "Gen", "int", "", "", "", ""],
        ["表头", "pokemon_data", "属性（中文）", "Type", "string", "", "", "", ""],
        ["表头", "pokemon_data", "属性（英文）", "TypeEn", "string", "", "", "", ""],
        ["表头", "pokemon_data", "生命值", "HP", "int", "", "", "", ""],
        ["表头", "pokemon_data", "攻击", "Attack", "int", "", "", "", ""],
        ["表头", "pokemon_data", "防御", "Defense", "int", "", "", "", ""],
        ["表头", "pokemon_data", "特攻", "SpAttack", "int", "", "", "", ""],
        ["表头", "pokemon_data", "特防", "SpDefense", "int", "", "", "", ""],
        ["表头", "pokemon_data", "速度", "Speed", "int", "", "", "", ""],
        ["表头", "pokemon_data", "身高", "Height", "float", "", "", "", ""],
        ["表头", "pokemon_data", "体重", "Weight", "float", "", "", "", ""],
        ["表头", "pokemon_data", "特性（英文）", "SpecialNature", "string", "", "", "", ""],
        ["表头", "pokemon_data", "特性（中文）", "ChineseSpecialNature", "string", "", "", "", ""],
        ["表头", "pokemon_data", "捕获率（网页）", "CaptureRate", "string", "", "", "", ""],
        ["表头", "pokemon_data", "孵化周期（网页）", "HatchCycle", "string", "", "", "", ""],
        ["表头", "pokemon_data", "性别比例（网页）", "GenderRatio", "string", "", "", "", ""],
        ["表头", "pokemon_data", "基础经验值（网页）", "BaseExp", "int", "", "", "", ""],
        ["表头", "pokemon_data", "HP基础点数（网页）", "BasePointHP", "int", "", "", "", ""],
        ["表头", "pokemon_data", "攻击基础点数（网页）", "BasePointAtk", "int", "", "", "", ""],
        ["表头", "pokemon_data", "防御基础点数（网页）", "BasePointDef", "int", "", "", "", ""],
        ["表头", "pokemon_data", "特攻基础点数（网页）", "BasePointSpAtk", "int", "", "", "", ""],
        ["表头", "pokemon_data", "特防基础点数（网页）", "BasePointSpDef", "int", "", "", "", ""],
        ["表头", "pokemon_data", "速度基础点数（网页）", "BasePointSpeed", "int", "", "", "", ""],
        ["表头", "pokemon_data", "进化条件（网页）", "EvolutionCondition", "string", "", "", "", ""],
        ["表头", "pokemon_data", "升级技能", "LevelUpMoves", "string", "", "", "", ""]
    ]
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save("C:\\Users\\90953\\Downloads\\my-pokemon-excel-designer\\pokemon_Type.xlsx")

def main():
    print("开始创建 Excel 文件...")
    # create_ability_index_xlsx()
    # create_item_index_xlsx()
    # create_move_index_xlsx()
    # create_pokemon_index_xlsx()

    # create_ability_type_xlsx()
    # create_item_type_xlsx()
    # create_move_type_xlsx()
    create_pokemon_type_xlsx()
    print("\n所有文件创建完成！")

if __name__ == "__main__":
    main()