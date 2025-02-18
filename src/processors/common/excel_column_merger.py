import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from typing import List

class ExcelColumnMerger:
    def __init__(self):
        self.output_dir = Path('output')
        
    def merge_columns(self, 
                     source_file: str, 
                     target_file: str, 
                     columns_to_merge: List[str],
                     name_column: str = '英文名称',
                     column_widths: dict = None):
        """
        将源表格的指定列合并到目标表格中
        
        参数:
            source_file: 源Excel文件路径
            target_file: 目标Excel文件路径
            columns_to_merge: 要合并的列名列表
            name_column: 用于匹配的列名（默认为'英文名称'）
            column_widths: 列宽度设置字典 {列名: 宽度}
        """
        print(f"读取源文件: {source_file}")
        df_source = pd.read_excel(source_file)
        
        print(f"读取目标文件: {target_file}")
        df_target = pd.read_excel(target_file)
        
        # 验证源文件中是否包含所有需要合并的列
        missing_columns = [col for col in columns_to_merge if col not in df_source.columns]
        if missing_columns:
            raise ValueError(f"源文件中缺少以下列: {missing_columns}")
            
        # 验证两个文件都包含匹配列
        if name_column not in df_source.columns or name_column not in df_target.columns:
            raise ValueError(f"两个文件都必须包含匹配列: {name_column}")
            
        print(f"\n开始合并列: {columns_to_merge}")
        
        # 创建列宽度配置
        if column_widths is None:
            column_widths = {col: 20 for col in columns_to_merge}  # 默认宽度为20
            
        # 对每个要合并的列进行处理
        for column in columns_to_merge:
            print(f"\n处理列: {column}")
            # 创建源数据的字典，用于快速查找
            source_dict = dict(zip(df_source[name_column], df_source[column]))
            
            # 如果目标表格没有该列，就创建新列
            if column not in df_target.columns:
                df_target[column] = None
                print(f"在目标表格中创建新列: {column}")
            
            # 更新数据
            matched_count = 0
            for idx, row in df_target.iterrows():
                name = row[name_column]
                if name in source_dict:
                    df_target.at[idx, column] = source_dict[name]
                    matched_count += 1
            
            print(f"匹配到 {matched_count} 行数据")
        
        # 使用openpyxl保存并设置列宽
        print("\n保存更新后的文件...")
        
        # 创建一个ExcelWriter对象
        with pd.ExcelWriter(target_file, engine='openpyxl') as writer:
            # 保存DataFrame
            df_target.to_excel(writer, index=False, sheet_name='Sheet1')
            
            # 获取worksheet
            worksheet = writer.sheets['Sheet1']
            
            # 设置列宽
            for idx, column in enumerate(df_target.columns, 1):
                if column in column_widths:
                    worksheet.column_dimensions[get_column_letter(idx)].width = column_widths[column]
        
        print("完成!")
        
def main():
    merger = ExcelColumnMerger()
    
    # 示例用法
    source_file = 'output/pokemon_web_info.xlsx'
    target_file = 'output/pokemon_data.xlsx'
    columns_to_merge = [
        '捕获率（网页）',
        '孵化周期（网页）',
        '性别比例（网页）', 
        '基础经验值（网页）',
        'HP基础点数（网页）',
        '攻击基础点数（网页）',
        '防御基础点数（网页）',
        '特攻基础点数（网页）',
        '特防基础点数（网页）',
        '速度基础点数（网页）'
        ]  # 要合并的列
    column_widths = {
        '捕获率（网页）': 20,
        '孵化周期（网页）': 25,
        '性别比例（网页）': 25,
        '基础经验值（网页）': 25,
        'HP基础点数（网页）': 25,
        '攻击基础点数（网页）': 25,
        '防御基础点数（网页）': 25,
        '特攻基础点数（网页）': 25,
        '特防基础点数（网页）': 25,
        '速度基础点数（网页）': 25
    }
    
    merger.merge_columns(
        source_file=source_file,
        target_file=target_file,
        columns_to_merge=columns_to_merge,
        column_widths=column_widths
    )

if __name__ == "__main__":
    main() 