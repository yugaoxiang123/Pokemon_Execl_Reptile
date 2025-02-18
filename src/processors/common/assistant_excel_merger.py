import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from typing import List

class AssistantExcelColumnMerger:
    def __init__(self):
        self.output_dir = Path('output')
        
    def merge_sheets(self, 
                    excel_file: str,
                    source_sheet: str,
                    target_sheet: str,
                    columns_to_merge: List[str],
                    name_column: str = '英文名称',
                    column_widths: dict = None):
        """
        将同一Excel文件中源表格的指定列合并到目标表格
        
        参数:
            excel_file: Excel文件路径
            source_sheet: 源表名称
            target_sheet: 目标表名称
            columns_to_merge: 要合并的列名列表
            name_column: 用于匹配的列名（默认为'英文名称'）
            column_widths: 列宽度设置字典 {列名: 宽度}
        """
        print(f"读取Excel文件: {excel_file}")
        with pd.ExcelFile(excel_file) as xls:
            df_source = pd.read_excel(xls, sheet_name=source_sheet)
            df_target = pd.read_excel(xls, sheet_name=target_sheet)
        
        # 验证源表中是否包含所有需要合并的列
        missing_columns = [col for col in columns_to_merge if col not in df_source.columns]
        if missing_columns:
            raise ValueError(f"源表中缺少以下列: {missing_columns}")
            
        # 验证两个表都包含匹配列
        if name_column not in df_source.columns or name_column not in df_target.columns:
            raise ValueError(f"两个表都必须包含匹配列: {name_column}")
            
        print(f"\n开始合并列: {columns_to_merge}")
        
        # 创建列宽度配置
        if column_widths is None:
            column_widths = {col: 20 for col in columns_to_merge}
            
        # 对每个要合并的列进行处理
        for column in columns_to_merge:
            print(f"\n处理列: {column}")
            source_dict = dict(zip(df_source[name_column], df_source[column]))
            
            if column not in df_target.columns:
                df_target[column] = None
                print(f"在目标表中创建新列: {column}")
            
            # 更新数据
            matched_count = 0
            for idx, row in df_target.iterrows():
                name = row[name_column]
                if name in source_dict:
                    df_target.at[idx, column] = source_dict[name]
                    matched_count += 1
            
            print(f"匹配到 {matched_count} 行数据")
        
        # 保存更新后的数据
        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_target.to_excel(writer, sheet_name=target_sheet, index=False)
            df_source.to_excel(writer, sheet_name=source_sheet, index=False)
            
            # 设置列宽
            worksheet = writer.sheets[target_sheet]
            for idx, column in enumerate(df_target.columns, 1):
                if column in column_widths:
                    worksheet.column_dimensions[get_column_letter(idx)].width = column_widths[column]
        
        print("完成!")

def main():
    merger = AssistantExcelColumnMerger()
    
    # 示例用法
    source_file = 'output/pokemon_data.xlsx'
    target_file = 'output/pokemon_data.xlsx'
    columns_to_merge = [
        '进化条件',
        '升级技能'
        ]  # 要合并的列
    column_widths = {
        '进化条件': 50,
        '升级技能': 150
    }
    
    merger.merge_sheets(
        excel_file=source_file,
        source_sheet='LevelUpMoves',
        target_sheet='Sheet1',
        columns_to_merge=columns_to_merge,
        column_widths=column_widths
    )

if __name__ == "__main__":
    main() 