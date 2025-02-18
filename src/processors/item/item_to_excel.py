import re
import json
import time
import pandas as pd
import requests
from bs4 import BeautifulSoup
from pathlib import Path
from tqdm import tqdm
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

class ItemProcessor:
    def __init__(self):
        self.item_cache_file = Path('json/item_translations.json')
        self.item_cache = self._load_cache()
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })

    def _load_cache(self) -> dict:
        """加载道具翻译缓存"""
        if self.item_cache_file.exists():
            return json.loads(self.item_cache_file.read_text(encoding='utf-8'))
        return {}

    def _save_cache(self):
        """保存道具翻译缓存"""
        self.item_cache_file.parent.mkdir(exist_ok=True)
        self.item_cache_file.write_text(
            json.dumps(self.item_cache, ensure_ascii=False, indent=2),
            encoding='utf-8'
        )

    def _get_item_translation(self, item_en: str) -> str:
        """获取道具的中文翻译"""
        if item_en in self.item_cache:
            return self.item_cache[item_en]

        try:
            url = f"https://wiki.52poke.com/wiki/{item_en}"
            response = self.session.get(url, timeout=10)
            soup = BeautifulSoup(response.text, 'html.parser')
            title = soup.find('h1', {'id': 'firstHeading'})
            if title:
                item_cn = title.text.split('（')[0].strip()
                self.item_cache[item_en] = item_cn
                self._save_cache()
                return item_cn
            time.sleep(1)
        except Exception as e:
            print(f"获取道具翻译失败 {item_en}: {e}")
        return item_en

    def parse_item_data(self, content: str) -> List[Dict]:
        """解析道具数据"""
        items = []
        
        # 使用正则表达式匹配每个道具对象
        pattern = r'(\w+):\s*{([^}]+)}'
        matches = list(re.finditer(pattern, content))

        print("解析道具数据...")
        for match in tqdm(matches, desc="处理道具"):
            try:
                item_id = match.group(1)
                data_block = match.group(2)
                
                # 提取所需字段
                name_match = re.search(r'name:\s*"([^"]+)"', data_block)
                num_match = re.search(r'num:\s*(\d+)', data_block)
                spritenum_match = re.search(r'spritenum:\s*(\d+)', data_block)
                
                if name_match:  # 只处理有名称的道具
                    item_data = {
                        'name_en': name_match.group(1),
                        'name_cn': self._get_item_translation(name_match.group(1)),
                        'num': int(num_match.group(1)) if num_match else None,
                        'spritenum': int(spritenum_match.group(1)) if spritenum_match else None
                    }
                    
                    # 获取中文翻译

                    items.append(item_data)

            except Exception as e:
                print(f"解析错误: {e}")
                continue

        return items

    def process_data(self, input_file: str, output_file: str):
        """处理数据主函数"""
        print("读取数据文件...")
        with open(input_file, 'r', encoding='utf-8') as file:
            content = file.read()

        # 解析数据
        items = self.parse_item_data(content)
        df = pd.DataFrame(items)

        # 重命名列
        column_rename = {
            'name_en': '道具名称（英文）',
            'name_cn': '道具名称（中文）',
            'num': '编号',
            'spritenum': '精灵图编号'
        }

        # 选择并重命名列
        df = df.rename(columns=column_rename)

        # 确保输出目录存在
        Path(output_file).parent.mkdir(exist_ok=True)

        # 使用 openpyxl 保存并设置列宽
        wb = Workbook()
        ws = wb.active

        # 写入表头
        headers = list(column_rename.values())
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # 写入数据
        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # 设置列宽
        column_widths = {
            '道具ID': 20,
            '道具名称（中文）': 25,
            '道具名称（英文）': 30,
            '编号': 10,
            '精灵图编号': 15
        }

        for col, header in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col)].width = column_widths[header]

        # 保存文件
        print(f"保存数据到 {output_file}")
        wb.save(output_file)

def main():
    processor = ItemProcessor()
    processor.process_data('data/items.ts', 'output/item_data.xlsx')

if __name__ == "__main__":
    main() 