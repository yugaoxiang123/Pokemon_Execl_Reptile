import re
import pandas as pd
from tqdm import tqdm
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class ItemGenExtractor:
    def __init__(self):
        self.data_file = Path('data/items.ts')
        self.excel_file = Path('output/item_data.xlsx')
        
    def _extract_gen_info(self) -> dict:
        """从items.ts文件中提取道具的世代信息"""
        gen_info = {}
        try:
            print("读取数据文件...")
            with open(self.data_file, 'r', encoding='utf-8') as file:
                content = file.read()
            
            # 移除 export const Items 声明
            content = re.sub(r'^export\s+const\s+Items\s*:\s*[^=]+=\s*', '', content.strip())
            content = content.rstrip(';')
            
            # 从后往前匹配
            # 先找到每个道具的结尾部分（包含gen和num的部分）
            end_pattern = r'gen:\s*(\d+)[^}]*?},'
            end_matches = list(re.finditer(end_pattern, content))
            
            print(f"找到 {len(end_matches)} 个可能的道具数据")
            
            for match in tqdm(end_matches, desc="处理道具"):
                # 从这个位置向前查找最近的name字段
                end_pos = match.end()
                gen = match.group(1)
                
                # 在当前位置之前的内容中查找最近的name字段
                prev_content = content[:end_pos]
                name_match = list(re.finditer(r'name:\s*"([^"]+)"', prev_content))
                
                if name_match:
                    # 取最后一个匹配（最近的name）
                    last_name_match = name_match[-1]
                    item_name = last_name_match.group(1)
                    # if len(name_match) > 1:
                    #     print(f"警告: 找到多个name字段 ({len(name_match)}个)")
                    #     for i, m in enumerate(name_match):
                    #         print(f"  {i+1}: {m.group(1)}")
                    
                    # 查找道具ID
                    id_section = prev_content[max(0, last_name_match.start() - 100):last_name_match.start()]
                    id_match = re.search(r'(\w+):\s*{[^{]*$', id_section)
                    item_id = id_match.group(1) if id_match else "unknown"
                    
                    # 查找编号
                    num_match = re.search(r'num:\s*(\d+)', match.group(0))
                    
                    gen_info[item_name] = int(gen)
                    print(f"成功: {item_name} (ID: {item_id}) - 世代 {gen}")
                    if num_match:
                        print(f"道具编号: {num_match.group(1)}")
            
            print(f"\n成功提取了 {len(gen_info)} 个道具的世代信息")
            if gen_info:
                print("\n示例数据:")
                sample_items = list(gen_info.items())[:5]
                for item, gen in sample_items:
                    print(f"{item}: 世代 {gen}")
                    
            return gen_info
            
        except Exception as e:
            print(f"提取世代信息失败: {e}")
            import traceback
            print(traceback.format_exc())
            return {}
    
    def update_excel_with_gen(self):
        """更新Excel文件，添加世代信息和编号"""
        try:
            print("读取世代信息...")
            gen_info = self._extract_gen_info()
            
            # 创建编号信息字典
            num_info = {}
            for match in re.finditer(r'name:\s*"([^"]+)".*?(?<!sprite)num:\s*(\d+)', self.data_file.read_text(encoding='utf-8'), re.DOTALL):
                item_name, num = match.groups()
                num_info[item_name] = int(num)
            
            print("读取Excel文件...")
            wb = load_workbook(self.excel_file)
            ws = wb.active
            
            # 找到需要的列
            header_row = list(ws.iter_rows(min_row=1, max_row=1))[0]
            name_en_col = None
            item_name_cn_col = None
            gen_col = None
            num_col = None
            
            for idx, cell in enumerate(header_row, 1):
                if cell.value == '道具名称（英文）':
                    name_en_col = idx
                elif cell.value == '道具名称（中文）':
                    item_name_cn_col = idx
                elif cell.value == '世代':
                    gen_col = idx
                elif cell.value == '编号':
                    num_col = idx
            
            if not name_en_col or not item_name_cn_col:
                print("找不到必要的列")
                return
            
            # 如果存在世代列，先删除
            # if gen_col:
            #     print("删除已存在的世代列...")
            #     ws.delete_cols(gen_col)
            
            # 在精灵图编号列后插入世代列
            # ws.insert_cols(item_name_cn_col + 1)
            # ws.cell(row=1, column=item_name_cn_col + 1, value='世代')
            
            # 设置世代列宽
            # ws.column_dimensions[get_column_letter(item_name_cn_col + 1)].width = 10
            
            # 更新世代信息和编号
            print("更新世代信息和编号...")
            for row in range(2, ws.max_row + 1):
                item_name = ws.cell(row=row, column=name_en_col).value
                if item_name:
                    # 更新世代
                    if item_name in gen_info:
                        ws.cell(row=row, column=gen_col, value=gen_info[item_name])
                    # 更新编号
                    if item_name in num_info and num_col:
                        ws.cell(row=row, column=num_col, value=num_info[item_name])
            
            # 保存文件
            print("保存更新后的Excel文件...")
            wb.save(self.excel_file)
            print("完成!")
            
        except Exception as e:
            print(f"更新Excel文件失败: {e}")

def main():
    extractor = ItemGenExtractor()
    extractor.update_excel_with_gen()

if __name__ == "__main__":
    main() 