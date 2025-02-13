import re
import json
import time
import pandas as pd
import requests
from bs4 import BeautifulSoup
from pathlib import Path
from tqdm import tqdm
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

class MoveProcessor:
    def __init__(self):
        self.move_cache_file = Path('json/move_translations.json')
        self.move_cache = self._load_cache()
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })

    def _load_cache(self) -> dict:
        """加载技能翻译缓存"""
        if self.move_cache_file.exists():
            return json.loads(self.move_cache_file.read_text(encoding='utf-8'))
        return {}

    def _save_cache(self):
        """保存技能翻译缓存"""
        self.move_cache_file.parent.mkdir(exist_ok=True)
        self.move_cache_file.write_text(
            json.dumps(self.move_cache, ensure_ascii=False, indent=2),
            encoding='utf-8'
        )

    def _get_move_translation(self, move_en: str) -> str:
        """获取技能的中文翻译"""
        if move_en in self.move_cache:
            return self.move_cache[move_en]

        try:
            url = f"https://wiki.52poke.com/wiki/{move_en}"
            response = self.session.get(url, timeout=10)
            soup = BeautifulSoup(response.text, 'html.parser')
            title = soup.find('h1', {'id': 'firstHeading'})
            if title:
                move_cn = title.text.split('（')[0].strip()
                self.move_cache[move_en] = move_cn
                self._save_cache()
                return move_cn
            time.sleep(1)
        except Exception as e:
            print(f"获取技能翻译失败 {move_en}: {e}")
        return move_en

    def parse_move_data(self, content: str) -> List[Dict]:
        """解析技能数据"""
        moves = []
        
        # 使用正则表达式匹配每个技能对象
        pattern = r'(\w+):\s*{([^}]+)}'
        matches = list(re.finditer(pattern, content))

        print("解析技能数据...")
        for match in tqdm(matches, desc="处理技能"):
            try:
                move_id = match.group(1)
                data_block = match.group(2)
                
                # 提取所需字段
                name_match = re.search(r'name:\s*"([^"]+)"', data_block)
                num_match = re.search(r'num:\s*(\d+)', data_block)
                accuracy_match = re.search(r'accuracy:\s*(\d+)', data_block)
                base_power_match = re.search(r'basePower:\s*(\d+)', data_block)
                pp_match = re.search(r'pp:\s*(\d+)', data_block)
                priority_match = re.search(r'priority:\s*(-?\d+)', data_block)
                category_match = re.search(r'category:\s*"([^"]+)"', data_block)
                
                if name_match:  # 只处理有名称的技能
                    move_data = {
                        'name_en': name_match.group(1),
                        'name_cn': self._get_move_translation(name_match.group(1)),
                        'num': int(num_match.group(1)) if num_match else None,
                        'accuracy': int(accuracy_match.group(1)) if accuracy_match else 'true',
                        'base_power': int(base_power_match.group(1)) if base_power_match else None,
                        'pp': int(pp_match.group(1)) if pp_match else None,
                        'priority': int(priority_match.group(1)) if priority_match else None,
                        'category': category_match.group(1) if category_match else None
                    }
                    moves.append(move_data)

            except Exception as e:
                print(f"解析错误: {e}")
                continue

        return moves

    def process_data(self, input_file: str, output_file: str):
        """处理数据主函数"""
        print("读取数据文件...")
        with open(input_file, 'r', encoding='utf-8') as file:
            content = file.read()

        # 解析数据
        moves = self.parse_move_data(content)
        df = pd.DataFrame(moves)

        # 重命名列
        column_rename = {
            'name_en': '技能名称（英文）',
            'name_cn': '技能名称（中文）',
            'num': '编号',
            'accuracy': '命中率',
            'base_power': '威力',
            'pp': 'PP值',
            'priority': '优先度',
            'category': '分类'
        }

        # 选择并重命名列
        df = df.rename(columns=column_rename)

        # 确保输出目录存在
        Path(output_file).parent.mkdir(exist_ok=True)

        # 使用 openpyxl 保存并设置列宽
        wb = Workbook()
        ws = wb.active

        # 写入表头
        headers = list(column_rename.values())
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # 写入数据
        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # 设置列宽
        column_widths = {
            '技能名称（中文）': 25,
            '技能名称（英文）': 30,
            '编号': 10,
            '命中率': 10,
            '威力': 10,
            'PP值': 10,
            '优先度': 10,
            '分类': 15
        }

        for col, header in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col)].width = column_widths[header]

        # 保存文件
        print(f"保存数据到 {output_file}")
        wb.save(output_file)

def main():
    processor = MoveProcessor()
    processor.process_data('data/moves.ts', 'output/move_data.xlsx')

if __name__ == "__main__":
    main() 