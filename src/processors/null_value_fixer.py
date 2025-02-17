import pandas as pd
from pathlib import Path
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class NullValueFixer:
    def __init__(self):
        self.output_dir = Path('output')
        
    def fix_null_values(self):
        """遍历所有Excel文件，将空值替换为'null'字符串"""
        try:
            # 获取所有xlsx文件
            excel_files = list(self.output_dir.glob('*.xlsx'))
            print(f"找到 {len(excel_files)} 个Excel文件")
            
            for excel_file in excel_files:
                print(f"\n处理文件: {excel_file.name}")
                
                # 读取所有sheet的数据和列宽
                sheets_data = {}
                with pd.ExcelFile(excel_file) as xls:
                    wb = load_workbook(excel_file)
                    
                    for sheet_name in xls.sheet_names:
                        print(f"\n处理表格: {sheet_name}")
                        
                        # 读取数据
                        df = pd.read_excel(xls, sheet_name=sheet_name)
                        
                        # 获取列宽
                        ws = wb[sheet_name]
                        column_widths = {get_column_letter(i+1): ws.column_dimensions[get_column_letter(i+1)].width 
                                       for i in range(len(df.columns)) 
                                       if get_column_letter(i+1) in ws.column_dimensions}
                        
                        # 替换空值
                        total_nulls = df.isna().sum().sum()
                        if total_nulls > 0:
                            print(f"发现 {total_nulls} 个空值")
                            df = df.fillna('null')
                            print(f"已将空值替换为'null'")
                        else:
                            print("未发现空值")
                            
                        sheets_data[sheet_name] = {
                            'data': df,
                            'column_widths': column_widths
                        }
                
                # 保存所有更新后的数据
                with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                    for sheet_name, sheet_info in sheets_data.items():
                        df = sheet_info['data']
                        column_widths = sheet_info['column_widths']
                        
                        # 写入数据
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        
                        # 设置列宽
                        worksheet = writer.sheets[sheet_name]
                        for col, width in column_widths.items():
                            worksheet.column_dimensions[col].width = width
                
        except Exception as e:
            print(f"处理失败: {e}")
            import traceback
            print(traceback.format_exc())

def main():
    fixer = NullValueFixer()
    fixer.fix_null_values()

if __name__ == "__main__":
    main() 