import json
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from tqdm import tqdm

class PokemonMovesExtractor:
    def __init__(self):
        self.excel_file = Path('output/pokemon_data.xlsx')
        self.output_dir = Path('json')
        self.output_dir.mkdir(exist_ok=True)
        
    def extract_moves(self):
        """从Excel文件中提取宝可梦名称和升级技能数据"""
        try:
            print("读取Excel文件...")
            wb = load_workbook(self.excel_file)
            ws = wb.active

            # 找到需要的列
            header_row = list(ws.iter_rows(min_row=1, max_row=1))[0]
            name_en_col = None
            moves_col = None
            
            for idx, cell in enumerate(header_row, 1):
                if cell.value == '英文名称':
                    name_en_col = idx
                elif cell.value == '升级技能':
                    moves_col = idx

            if not name_en_col or not moves_col:
                print("找不到必要的列")
                return

            # 提取数据
            pokemon_moves = {}
            print("提取数据中...")
            for row in tqdm(range(2, ws.max_row + 1)):
                pokemon_name = ws.cell(row=row, column=name_en_col).value
                moves = ws.cell(row=row, column=moves_col).value
                
                if pokemon_name and moves:
                    pokemon_moves[pokemon_name] = moves

            # 保存为JSON文件
            output_file = self.output_dir / 'pokemon_moves.json'
            print(f"保存数据到 {output_file}")
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(pokemon_moves, f, ensure_ascii=False, indent=2)

            print(f"完成! 共处理了 {len(pokemon_moves)} 个宝可梦的数据")
            
            # 显示示例数据
            print("\n数据示例:")
            sample_items = list(pokemon_moves.items())[:3]
            for name, moves in sample_items:
                print(f"{name}: {moves[:100]}...")
            
        except Exception as e:
            print(f"处理数据失败: {e}")
            import traceback
            print(traceback.format_exc())

def main():
    extractor = PokemonMovesExtractor()
    extractor.extract_moves()

if __name__ == "__main__":
    main() 