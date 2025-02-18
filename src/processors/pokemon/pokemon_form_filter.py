import pandas as pd
import re
from pathlib import Path
from openpyxl import load_workbook

class PokemonFormFilter:
    def __init__(self):
        self.excel_file = Path('output/pokemon_data.xlsx')
        
    def filter_forms(self):
        """过滤宝可梦形态，只保留基础形态和超级形态"""
        try:
            print("读取Excel文件...")
            wb = load_workbook(self.excel_file)
            ws = wb.active

            # 找到中文名称列
            header_row = list(ws.iter_rows(min_row=1, max_row=1))[0]
            name_cn_col = None
            
            for idx, cell in enumerate(header_row, 1):
                if cell.value == '中文名称':
                    name_cn_col = idx
                    break

            if not name_cn_col:
                print("找不到中文名称列")
                return

            # 从后往前遍历，这样删除行时不会影响索引
            rows_to_delete = []
            for row in range(ws.max_row, 1, -1):
                name = ws.cell(row=row, column=name_cn_col).value
                if name:
                    # 检查是否有括号
                    if '(' in name or '（' in name:
                        # 统一括号格式
                        name = name.replace('（', '(').replace('）', ')')
                        # 提取括号中的内容
                        form = re.search(r'\((.*?)\)', name)
                        if form:
                            form_name = form.group(1)
                            # 如果不是"超级"形态，则标记为删除
                            if form_name != '超级':
                                rows_to_delete.append(row)
                                print(f"将删除: {name}")

            # 删除标记的行
            print(f"\n开始删除 {len(rows_to_delete)} 个非基础/超级形态...")
            for row in rows_to_delete:
                ws.delete_rows(row)

            # 保存文件
            print("保存更新后的Excel文件...")
            wb.save(self.excel_file)
            print(f"完成! 共删除了 {len(rows_to_delete)} 行数据")
            
        except Exception as e:
            print(f"处理Excel文件失败: {e}")
            import traceback
            print(traceback.format_exc())

def main():
    filter = PokemonFormFilter()
    filter.filter_forms()

if __name__ == "__main__":
    main() 