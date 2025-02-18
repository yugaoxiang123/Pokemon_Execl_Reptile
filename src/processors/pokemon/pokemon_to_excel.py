import re
import time
import json
import pandas as pd
import requests
from bs4 import BeautifulSoup
from typing import Dict, Any
from tqdm import tqdm
from pathlib import Path
from utils import FORM_TRANSLATIONS, SPECIAL_FORM_NAMES
from openpyxl import Workbook

# 属性中英文对照
TYPE_TRANSLATIONS = {
    'Normal': '一般', 'Fire': '火', 'Water': '水', 'Electric': '电',
    'Grass': '草', 'Ice': '冰', 'Fighting': '格斗', 'Poison': '毒',
    'Ground': '地面', 'Flying': '飞行', 'Psychic': '超能力', 'Bug': '虫',
    'Rock': '岩石', 'Ghost': '幽灵', 'Dragon': '龙', 'Dark': '恶',
    'Steel': '钢', 'Fairy': '妖精'
}

class PokemonDataProcessor:
    def __init__(self):
        self.cache_file = Path('json/pokemon_name_cache.json')
        self.name_cache = self._load_cache()
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
        self.form_translations = FORM_TRANSLATIONS
        self.special_form_names = SPECIAL_FORM_NAMES
        self.ability_cache_file = Path('json/ability_translations.json')
        self.ability_cache = self._load_ability_cache()

    def _load_cache(self) -> dict:
        """加载名称缓存"""
        if self.cache_file.exists():
            return json.loads(self.cache_file.read_text(encoding='utf-8'))
        return {}

    def _save_cache(self):
        """保存名称缓存"""
        self.cache_file.write_text(json.dumps(self.name_cache, ensure_ascii=False, indent=2), 
                                 encoding='utf-8')

    def _load_ability_cache(self) -> dict:
        """加载特性翻译缓存"""
        if self.ability_cache_file.exists():
            return json.loads(self.ability_cache_file.read_text(encoding='utf-8'))
        return {}

    def _save_ability_cache(self):
        """保存特性翻译缓存"""
        self.ability_cache_file.write_text(
            json.dumps(self.ability_cache, ensure_ascii=False, indent=2),
            encoding='utf-8'
        )

    def _get_ability_translation(self, ability_en: str) -> str:
        """获取特性的中文翻译"""
        if ability_en in self.ability_cache:
            return self.ability_cache[ability_en]

        try:
            url = f"https://wiki.52poke.com/wiki/{ability_en}"
            response = self.session.get(url, timeout=10)
            soup = BeautifulSoup(response.text, 'html.parser')
            title = soup.find('h1', {'id': 'firstHeading'})
            if title:
                ability_cn = title.text.split('（')[0].strip()
                self.ability_cache[ability_en] = ability_cn
                self._save_ability_cache()
                return ability_cn
            time.sleep(1)
        except Exception as e:
            print(f"获取特性翻译失败 {ability_en}: {e}")
        return ability_en

    def parse_pokemon_data(self, content: str) -> list[Dict[str, Any]]:
        """解析宝可梦数据文件"""
        pokemon_list = []
        pattern = r'(\w+):\s*{((?:[^{}]|{[^{}]*})*)}'
        matches = list(re.finditer(pattern, content))  # 转换为列表以便使用tqdm
        
        print("解析宝可梦数据和翻译特性...")
        for match in tqdm(matches, desc="处理宝可梦"):
            try:
                # 初始化所有必需的字段
                pokemon_data = {
                    'id': None,
                    'name_en': None,
                    'types_en': None,
                    'types_cn': None,
                    'hp': None,
                    'atk': None,
                    'def': None,
                    'spa': None,
                    'spd': None,
                    'spe': None,
                    'height': None,
                    'weight': None,
                    'ability': None,
                    'abilities_en': None,
                    'abilities_cn': None
                }
                
                data_block = match.group(2)
                
                # 提取数字ID
                num_match = re.search(r'num:\s*([-\d]+)', data_block)
                if not num_match:
                    continue
                pokemon_data['id'] = num_match.group(1)
                
                # 提取名称
                name_match = re.search(r'name:\s*"([^"]+)"', data_block)
                if name_match:
                    pokemon_data['name_en'] = name_match.group(1)
                
                # 提取类型
                types_match = re.search(r'types:\s*\[(.*?)\]', data_block)
                if types_match:
                    types = types_match.group(1)
                    types_list = [t.strip(' "\'') for t in types.split(',')]
                    pokemon_data['types_en'] = ', '.join(types_list)
                    pokemon_data['types_cn'] = ', '.join(TYPE_TRANSLATIONS.get(t, t) for t in types_list)
                
                # 提取基础属性
                stats_match = re.search(r'baseStats:\s*{([^}]+)}', data_block)
                if stats_match:
                    stats_text = stats_match.group(1)
                    # 使用正则表达式分别匹配每个属性
                    for stat in ['hp', 'atk', 'def', 'spa', 'spd', 'spe']:
                        stat_match = re.search(rf'{stat}:\s*(\d+)', stats_text)
                        if stat_match:
                            pokemon_data[stat] = int(stat_match.group(1))
                
                # 提取身高体重
                height_match = re.search(r'heightm:\s*([\d.]+)', data_block)
                if height_match:
                    pokemon_data['height'] = float(height_match.group(1))
                
                weight_match = re.search(r'weightkg:\s*([\d.]+)', data_block)
                if weight_match:
                    pokemon_data['weight'] = float(weight_match.group(1))
                
                # 提取特性
                abilities_match = re.search(r'abilities:\s*{([^}]+)}', data_block)
                if abilities_match:
                    abilities_text = abilities_match.group(1)
                    # 使用正则表达式匹配键值对
                    ability_pairs = re.findall(r'(\w+):\s*"([^"]+)"', abilities_text)
                    
                    # 翻译特性（不使用列表推导式，以便显示进度）
                    abilities_en = []
                    abilities_cn = []
                    for k, v in ability_pairs:
                        abilities_en.append(f"{k}:{v}")
                        cn_ability = self._get_ability_translation(v)
                        abilities_cn.append(f"{k}:{cn_ability}")
                    
                    pokemon_data['abilities_en'] = ', '.join(abilities_en)
                    pokemon_data['abilities_cn'] = ', '.join(abilities_cn)
                
                pokemon_list.append(pokemon_data)
                
            except Exception as e:
                print(f"解析错误: {e}")
                print(f"出错的数据块: {data_block[:100]}...")
                continue
        
        return pokemon_list

    def get_chinese_name(self, name_en: str, max_retries: int = 3) -> str:
        """获取中文名称（带重试机制）"""
        if name_en in self.name_cache:
            return self.name_cache[name_en]
        # 特殊形态的中文翻译对照
        form_translations = self.form_translations
        
        # 特殊处理某些完整形态名称
        special_form_names = self.special_form_names
        
        # 检查是否有特殊完整形态名称
        if name_en in special_form_names:
            return special_form_names[name_en]
        
        # 处理特殊形态的宝可梦名称
        if '-' in name_en:
            base_name, *forms = name_en.split('-')
            # 获取基础形态的中文名
            base_chinese = self.get_chinese_name(base_name)
            
            # 翻译形态
            translated_forms = []
            current_form = '-'.join(forms)  # 处理多段形态名称
            
            # 先尝试整体匹配
            if current_form in form_translations:
                translated_forms.append(form_translations[current_form])
            else:
                # 否则逐个翻译
                for form in forms:
                    if form in form_translations:
                        translated_forms.append(form_translations[form])
                    else:
                        translated_forms.append(form)
            
            # 组合完整的中文名称，将形态用括号括起来
            chinese_name = f"{base_chinese}（{''.join(translated_forms)}）"
            self.name_cache[name_en] = chinese_name
            return chinese_name
        
        # 原有的网页查询逻辑
        url = f"https://wiki.52poke.com/wiki/{name_en}"
        
        for attempt in range(max_retries):
            try:
                response = self.session.get(url, timeout=10)
                soup = BeautifulSoup(response.text, 'html.parser')
                title = soup.find('h1', {'id': 'firstHeading'})
                
                if title:
                    chinese_name = title.text.split('(')[0].strip()
                    self.name_cache[name_en] = chinese_name
                    return chinese_name
                    
                time.sleep(1)  # 请求间隔
                
            except Exception as e:
                if attempt == max_retries - 1:
                    print(f"获取{name_en}的中文名称失败: {e}")
                time.sleep(2)  # 失败后等待更长时间
                
        return "未知"

    def process_data(self, input_file: str, output_file: str):
        """处理数据主函数"""
        print("读取数据文件...")
        with open(input_file, 'r', encoding='utf-8') as file:
            content = file.read()
        
        print("解析宝可梦数据...")
        pokemon_list = self.parse_pokemon_data(content)
        df = pd.DataFrame(pokemon_list)
        
        print("获取中文名称...")
        tqdm.pandas()
        df['name_cn'] = df['name_en'].progress_apply(self.get_chinese_name)
        
        # 检查images目录是否存在
        image_dir = Path('images')
        if not image_dir.exists():
            print("创建images目录...")
            image_dir.mkdir(exist_ok=True)
        
        # 检查已下载的图片
        existing_images = {f.stem: f for f in image_dir.glob('*.png')}
        
        # # 下载缺失的图片：暂时注释掉 因为有些图片网上无法找到
        # missing_pokemon = [name for name in df['name_en'] if name not in existing_images]
        # if missing_pokemon:
        #     print(f"需要下载 {len(missing_pokemon)} 个宝可梦图片...")
        #     image_downloader = PokemonImageDownloader()
        #     df['image_path'] = df['name_en'].progress_apply(image_downloader.download_image)
        # else:
        #     print("所有图片已下载，跳过下载步骤...")
        #     # 直接设置图片路径
        #     df['image_path'] = df['name_en'].apply(lambda x: str(existing_images[x]))

        # 使用 get 方法，提供默认值
        df['image_path'] = df['name_en'].apply(lambda x: str(existing_images.get(x, '')))
        # 保存缓存
        self._save_cache()
        
        # 检查数据框中是否存在所需的列
        required_columns = ['id', 'name_cn', 'name_en', 'types_cn', 'types_en',
                           'hp', 'atk', 'def', 'spa', 'spd', 'spe',
                           'height', 'weight', 'abilities_en', 'abilities_cn', 'image_path']
        
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            print("缺失的列:", missing_columns)
        
        # 确保所有必需的列都存在
        for col in required_columns:
            if col not in df.columns:
                df[col] = None
        
        # 修改列名和顺序
        columns = [
            'id', 'name_cn', 'name_en', 
            'types_cn', 'types_en',
            'hp', 'atk', 'def', 'spa', 'spd', 'spe',
            'height', 'weight', 'abilities_en', 'abilities_cn',
            'image_path'  # 保留图片路径列
        ]
        
        # 重命名列（全部改为中文）
        column_rename = {
            'id': '编号',
            'name_cn': '中文名称',
            'name_en': '英文名称',
            'types_cn': '属性（中文）',
            'types_en': '属性（英文）',
            'hp': '生命值',
            'atk': '攻击',
            'def': '防御',
            'spa': '特攻',
            'spd': '特防',
            'spe': '速度',
            'height': '身高',
            'weight': '体重',
            'abilities_en': '特性（英文）',
            'abilities_cn': '特性（中文）',
            'image_path': '图鉴路径'
        }
        
        # 选择并重命名列
        df = df[columns].rename(columns=column_rename)

        # 设置样式并保存
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Pokemon Data'
        
        # 先写入表头
        headers = df.columns.tolist()
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)
        
        # 写入数据
        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)
        
        # 调整列宽
        for column in worksheet.columns:
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # 保存文件
        try:
            workbook.save(output_file)
            print(f"数据已保存到 {output_file}")
        except Exception as e:
            print(f"保存Excel文件失败: {e}")

def main():
    processor = PokemonDataProcessor()
    processor.process_data('data/pokedex.ts', 'output/pokemon_data.xlsx')

    # 确保输出目录存在
    Path('output').mkdir(exist_ok=True)

if __name__ == "__main__":
    main() 