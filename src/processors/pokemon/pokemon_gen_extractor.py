import json
import pandas as pd
from tqdm import tqdm
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class PokemonGenExtractor:
    def __init__(self):
        self.json_file = Path('json/pokemons_by_gen.json')
        self.excel_file = Path('output/pokemon_data.xlsx')
        
    def get_generation_by_pokemon(self, pokemon_name: str) -> int:
        """根据宝可梦名称获取其世代"""
        try:
            # 读取 JSON 数据
            with open(self.json_file, 'r', encoding='utf-8') as f:
                pokemon_by_gen = json.load(f)

            # 处理可能的形态变化（去除-Mega, -Gmax等后缀）
            # base_name = pokemon_name.split('-')[0]

            # 遍历每个世代的宝可梦
            for gen, pokemons in pokemon_by_gen.items():
                # if base_name in pokemons:
                #     return int(gen)
                # 检查完整名称（包含形态）
                if pokemon_name in pokemons:
                    return int(gen)

            return None  # 如果没有找到，返回 None
            
        except Exception as e:
            print(f"获取世代信息时出错: {e}")
            return None

    def update_excel_with_gen(self):
        """更新Excel文件，添加世代信息"""
        try:
            print("读取Excel文件...")
            wb = load_workbook(self.excel_file)
            ws = wb.active

            # 找到需要的列
            header_row = list(ws.iter_rows(min_row=1, max_row=1))[0]
            name_en_col = None
            pokemon_name_cn_col = None
            gen_col = None
            
            for idx, cell in enumerate(header_row, 1):
                if cell.value == '英文名称':
                    name_en_col = idx
                elif cell.value == '中文名称':
                    pokemon_name_cn_col = idx
                elif cell.value == '世代':
                    gen_col = idx

            if not name_en_col:
                print("找不到宝可梦英文名称列")
                return
            if not pokemon_name_cn_col:
                print("找不到宝可梦中文名称列")
                return

            # 如果没有世代列，创建一个
            if not gen_col:
                ws.insert_cols(pokemon_name_cn_col + 1)
                ws.cell(row=1, column=pokemon_name_cn_col + 1, value='世代')
                gen_col = pokemon_name_cn_col + 1

            # 设置世代列宽度
            ws.column_dimensions[get_column_letter(gen_col)].width = 10

            # 遍历所有宝可梦
            print("更新世代信息...")
            for row in tqdm(range(2, ws.max_row + 1)):
                pokemon_en = ws.cell(row=row, column=name_en_col).value
                if pokemon_en:
                    gen = self.get_generation_by_pokemon(pokemon_en)
                    if gen is None:
                        gen = -99  # 使用-99表示未找到世代信息
                    ws.cell(row=row, column=gen_col, value=gen)

            # 保存更新后的文件
            print("保存更新后的Excel文件...")
            wb.save(self.excel_file)
            print("完成!")
            
        except Exception as e:
            print(f"更新Excel文件失败: {e}")
            import traceback
            print(traceback.format_exc())

def main():
    extractor = PokemonGenExtractor()
    extractor.update_excel_with_gen()

if __name__ == "__main__":
    main() 