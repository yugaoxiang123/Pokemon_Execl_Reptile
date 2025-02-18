import json
import pandas as pd
from tqdm import tqdm
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class MoveGenExtractor:
    def __init__(self):
        self.json_file = Path('json/moves_by_gen.json')
        self.excel_file = Path('output/move_data.xlsx')
        
    def get_generation_by_move(self, move_name):
        # 读取 JSON 数据
        with open(self.json_file, 'r', encoding='utf-8') as f:
            move_by_gen = json.load(f)

        # 遍历每个世代的能力
        for gen, abilities in move_by_gen.items():
            if move_name in abilities:
                return int(gen)  # 返回世代数字

        return None  # 如果没有找到，返回 None
    def update_excel_with_gen(self):
        """更新Excel文件，添加世代信息和编号"""
        try:
            print("读取Excel文件...")
            wb = load_workbook(self.excel_file)
            ws = wb.active

            # 找到英文名称列
            header_row = list(ws.iter_rows(min_row=1, max_row=1))[0]
            name_en_col = None
            move_name_cn_col = None
            gen_col = None
            for idx, cell in enumerate(header_row, 1):
                if cell.value == '技能名称（英文）':
                    name_en_col = idx
                elif cell.value == '技能名称（中文）':
                    move_name_cn_col = idx
                elif cell.value == '世代':
                    gen_col = idx

            if not name_en_col:
                print("找不到特性英文名称列")
                return
            if not move_name_cn_col:
                print("找不到特性中文名称列")
                return
            if not gen_col:
                ws.insert_cols(move_name_cn_col+1)
                ws.cell(row=1, column=move_name_cn_col+1, value='世代')
                gen_col = move_name_cn_col+1

            # 设置描述列宽度
            ws.column_dimensions[get_column_letter(gen_col)].width = 10

            # 遍历所有特性
            for row in tqdm(range(2, ws.max_row + 1)):
                move_en = ws.cell(row=row, column=name_en_col).value
                if move_en:
                    tmp_gen_index = self.get_generation_by_move(move_en)
                    if not tmp_gen_index:
                        tmp_gen_index = -99
                    ws.cell(row=row, column=gen_col, value=tmp_gen_index)

            # 保存更新后的文件
            print("保存更新后的Excel文件...")
            wb.save(self.excel_file)
        except Exception as e:
            print(f"更新Excel文件失败: {e}")

def main():
    extractor = MoveGenExtractor()
    extractor.update_excel_with_gen()

if __name__ == "__main__":
    main() 