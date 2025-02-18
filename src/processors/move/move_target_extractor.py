import re
import pandas as pd
from tqdm import tqdm
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class MoveTargetExtractor:
    def __init__(self):
        self.excel_file = "output/move_data.xlsx"
        self.data_file = "data/moves.ts"

    def extract_move_target(self) -> dict[str, str]:
        print("读取数据文件...")
        with open(self.data_file, 'r', encoding='utf-8') as file:
            content = file.read()
        
        # 移除 export const Items 声明
        content = re.sub(r'^export\s+const\s+Items\s*:\s*[^=]+=\s*', '', content.strip())
        content = content.rstrip(';')
        
        # 创建编号信息字典
        target_info = {}
        matches = list(re.finditer(r'name:\s*"([^"]+)".*?target:\s*"([^"]+)"', content, re.DOTALL))
        for match in tqdm (matches, desc="处理数据"):
            target_name, target = match.groups()
            target_info[target_name] = target.strip()
        return target_info

    def update_excel_with_target(self):
        """更新Excel文件，添加世代信息和编号"""
        try:
            target_info = self.extract_move_target()
            print("读取Excel文件...")
            wb = load_workbook(self.excel_file)
            ws = wb.active
            
            # 找到需要的列
            header_row = list(ws.iter_rows(min_row=1, max_row=1))[0]
            name_en_col = None
            priority_col = None
            target_col = None
            
            for idx, cell in enumerate(header_row, 1):
                if cell.value == '技能名称（英文）':
                    name_en_col = idx
                elif cell.value == '优先度':
                    priority_col = idx
                elif cell.value == '技能目标':
                    target_col = idx
            
            if not name_en_col or not priority_col:
                print("找不到必要的列")
                return
            
            if not target_col:
                target_col = priority_col + 1
                ws.insert_cols(target_col)
                ws.cell(row=1, column=target_col, value='技能目标')
                print("找不到技能目标列，已插入")

            print("更新世代信息和编号...")
            for row in range(2, ws.max_row + 1):
                target_name = ws.cell(row=row, column=name_en_col).value
                if target_name:
                    # 更新技能目标
                    if target_name in target_info and target_col:
                        ws.cell(row=row, column=target_col, value=target_info[target_name])
            
            # 保存文件
            print("保存更新后的Excel文件...")
            wb.save(self.excel_file)
            print("完成!")
            
        except Exception as e:
            print(f"更新Excel文件失败: {e}")
def main():
    extractor = MoveTargetExtractor()
    extractor.update_excel_with_target()

if __name__ == "__main__":
    main()

