import json
import time
import requests
from bs4 import BeautifulSoup, Tag
from pathlib import Path
from tqdm import tqdm
from typing import Optional, Tuple
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class MoveDescriptionScraper:
    def __init__(self):
        self.description_cache_file = Path('json/move_descriptions.json')
        self.description_cache = self._load_cache()
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })

    def _load_cache(self) -> dict:
        """加载技能描述缓存"""
        if self.description_cache_file.exists():
            return json.loads(self.description_cache_file.read_text(encoding='utf-8'))
        return {}

    def _save_cache(self):
        """保存技能描述缓存"""
        self.description_cache_file.write_text(
            json.dumps(self.description_cache, ensure_ascii=False, indent=2),
            encoding='utf-8'
        )

    def get_move_description(self, move_en: str) -> str:
        """获取技能的中文描述"""
        if move_en in self.description_cache:
            return self.description_cache[move_en]

        try:
            url = f"https://wiki.52poke.com/wiki/{move_en}"
            response = self.session.get(url, timeout=10)
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # 找到所有tbody标签
            tbodies = soup.find_all('tbody')
            
            # 遍历所有tbody，找到包含技能描述的那个
            for tbody in tbodies:
                trs = tbody.find_all('tr')
                if len(trs) > 1:  # 确保至少有两行
                    # 获取第二行的td标签
                    td = trs[1].find('td', {'class': 'roundy'})  # 技能描述通常有roundy类
                    if td and td.get('style') and 'font-size:smaller' in td.get('style'):
                        description = td.get_text(strip=True)
                        if description:
                            self.description_cache[move_en] = description
                            self._save_cache()
                            return description
            
            time.sleep(1)
        except Exception as e:
            print(f"获取技能描述失败 {move_en}: {e}")
        return ""

    def update_excel_with_descriptions(self, excel_file: str):
        """更新Excel文件，添加描述列"""
        print("读取Excel文件...")
        wb = load_workbook(excel_file)
        ws = wb.active

        # 找到英文名称列
        header_row = list(ws.iter_rows(min_row=1, max_row=1))[0]
        name_en_col = None
        for idx, cell in enumerate(header_row, 1):
            if cell.value == '技能名称（英文）':
                name_en_col = idx
                break

        if not name_en_col:
            print("找不到技能英文名称列")
            return

        desc_col = None
        for idx, cell in enumerate(header_row, 1):
            if cell.value == '技能描述':
                desc_col = idx
                break

        if not desc_col:
            # 添加描述列标题
            desc_col = len(list(ws.columns)) + 1
            ws.cell(row=1, column=desc_col, value='技能描述')

        # 设置描述列宽度
        ws.column_dimensions[get_column_letter(desc_col)].width = 100

        print("获取技能描述...")
        # 遍历所有技能
        for row in tqdm(range(2, ws.max_row + 1)):
            move_en = ws.cell(row=row, column=name_en_col).value
            if move_en:
                description = self.get_move_description(move_en)
                ws.cell(row=row, column=desc_col, value=description)

        # 保存更新后的文件
        print("保存更新后的Excel文件...")
        wb.save(excel_file)

def main():
    scraper = MoveDescriptionScraper()
    scraper.update_excel_with_descriptions('output/move_data.xlsx')

if __name__ == "__main__":
    main() 