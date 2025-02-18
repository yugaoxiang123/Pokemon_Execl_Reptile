import json
import time
import requests
from bs4 import BeautifulSoup, Tag
from pathlib import Path
from tqdm import tqdm
from typing import Optional, Tuple
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class ItemDescriptionScraper:
    def __init__(self):
        self.description_cache_file = Path('json/item_descriptions.json')
        self.description_cache = self._load_cache()
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })

    def _load_cache(self) -> dict:
        """加载道具描述缓存"""
        if self.description_cache_file.exists():
            return json.loads(self.description_cache_file.read_text(encoding='utf-8'))
        return {}

    def _save_cache(self):
        """保存道具描述缓存"""
        self.description_cache_file.parent.mkdir(exist_ok=True)
        self.description_cache_file.write_text(
            json.dumps(self.description_cache, ensure_ascii=False, indent=2),
            encoding='utf-8'
        )

    def _find_boundary_tags(self, soup: BeautifulSoup) -> Tuple[Optional[Tag], Optional[Tag]]:
        """找到描述的开始和结束标签"""
        # 尝试找开始标签
        start_tag = None
        for headline_id in ["效果", "游戏中", "使用效果"]:
            start_tag = soup.find('span', {'class': 'mw-headline', 'id': headline_id})
            if start_tag:
                start_tag = start_tag.find_parent(['h2', 'h3'])  # 同时支持h2和h3标签
                break

        if not start_tag:
            return None, None

        # 尝试找结束标签
        end_tag = None
        for headline_id in ["效果变更", "获得方式", "包包信息"]:
            end_tag = soup.find('span', {'class': 'mw-headline', 'id': headline_id})
            if end_tag:
                end_tag = end_tag.find_parent(['h2', 'h3'])  # 同时支持h2和h3标签
                break

        return start_tag, end_tag

    def get_item_description(self, item_en: str) -> str:
        """获取道具的中文描述"""
        if item_en in self.description_cache:
            description = self.description_cache[item_en]
            # 去除"获得方式"及其后面的内容
            if "获得方式" in description:
                description = description.split("获得方式")[0].strip() 
            self.description_cache[item_en] = description
            self._save_cache()
            return self.description_cache[item_en]

        try:
            url = f"https://wiki.52poke.com/wiki/{item_en}"
            response = self.session.get(url, timeout=10)
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # 找到边界标签
            start_tag, end_tag = self._find_boundary_tags(soup)
            if not start_tag:
                return ""

            # 收集描述文本
            description_parts = []
            current = start_tag.find_next_sibling()
            
            while current:
                # 检查是否到达结束标签
                if current.name in ['h2', 'h3']:
                    headline = current.find('span', {'class': 'mw-headline'})
                    if headline and headline.get('id') in ["效果变更", "获得方式", "包包信息"]:
                        break
                
                text = current.get_text(strip=True)
                if text:
                    description_parts.append(text)
                current = current.find_next_sibling()

            description = ' '.join(description_parts)
            
            # 去除"获得方式"及其后面的内容
            if "获得方式" in description:
                description = description.split("获得方式")[0].strip()
                
            if description:
                self.description_cache[item_en] = description
                self._save_cache()
                return description
            
            time.sleep(1)
        except Exception as e:
            print(f"获取道具描述失败 {item_en}: {e}")
        return ""

    def update_excel_with_descriptions(self, excel_file: str):
        """更新Excel文件，添加描述列"""
        print("读取Excel文件...")
        wb = load_workbook(excel_file)
        ws = wb.active

        # 找到英文名称列
        header_row = list(ws.iter_rows(min_row=1, max_row=1))[0]
        name_en_col = None
        for idx, cell in enumerate(header_row, 1):
            if cell.value == '道具名称（英文）':
                name_en_col = idx
                break

        if not name_en_col:
            print("找不到道具英文名称列")
            return

        desc_col = None
        for idx, cell in enumerate(header_row, 1):
            if cell.value == '道具描述':
                desc_col = idx
                break

        if not desc_col:
            # 添加描述列标题
            desc_col = len(list(ws.columns)) + 1
            ws.cell(row=1, column=desc_col, value='道具描述')

        # 设置描述列宽度
        ws.column_dimensions[get_column_letter(desc_col)].width = 100

        print("获取道具描述...")
        # 遍历所有道具
        for row in tqdm(range(2, ws.max_row + 1)):
            item_en = ws.cell(row=row, column=name_en_col).value
            if item_en:
                description = self.get_item_description(item_en)
                ws.cell(row=row, column=desc_col, value=description)

        # 保存更新后的文件
        print("保存更新后的Excel文件...")
        wb.save(excel_file)

def main():
    scraper = ItemDescriptionScraper()
    scraper.update_excel_with_descriptions('output/item_data.xlsx')

if __name__ == "__main__":
    main() 