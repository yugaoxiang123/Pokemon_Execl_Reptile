import json
import time
import requests
from bs4 import BeautifulSoup
from pathlib import Path
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class AbilityDescriptionScraper:
    def __init__(self):
        self.description_cache_file = Path('json/ability_descriptions.json')
        self.description_cache = self._load_cache()
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })

    def _load_cache(self) -> dict:
        """加载特性描述缓存"""
        if self.description_cache_file.exists():
            return json.loads(self.description_cache_file.read_text(encoding='utf-8'))
        return {}

    def _save_cache(self):
        """保存特性描述缓存"""
        self.description_cache_file.write_text(
            json.dumps(self.description_cache, ensure_ascii=False, indent=2),
            encoding='utf-8'
        )

    def get_ability_description(self, ability_en: str) -> str:
        """获取特性的中文描述"""
        if ability_en in self.description_cache:
            return self.description_cache[ability_en]

        try:
            url = f"https://wiki.52poke.com/wiki/{ability_en}"
            response = self.session.get(url, timeout=10)
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # 查找描述文本
            description_cell = soup.find('td', {'class': 'roundybottom-6 bgwhite', 'colspan': '2'})
            if description_cell:
                description = description_cell.text.strip()
                self.description_cache[ability_en] = description
                self._save_cache()
                return description
            time.sleep(1)
        except Exception as e:
            print(f"获取特性描述失败 {ability_en}: {e}")
        return ""

    def update_excel_with_descriptions(self, excel_file: str):
        """更新Excel文件，添加描述列"""
        print("读取Excel文件...")
        wb = load_workbook(excel_file)
        ws = wb.active

        # 找到英文名称列
        header_row = list(ws.iter_rows(min_row=1, max_row=1))[0]
        name_en_col = None
        for idx, cell in enumerate(header_row, 1):
            if cell.value == '特性名称（英文）':
                name_en_col = idx
                break

        if not name_en_col:
            print("找不到特性英文名称列")
            return

        # 添加描述列标题
        desc_col = len(list(ws.columns)) + 1
        ws.cell(row=1, column=desc_col, value='特性描述')

        # 设置描述列宽度
        ws.column_dimensions[get_column_letter(desc_col)].width = 50

        print("获取特性描述...")
        # 遍历所有特性
        for row in tqdm(range(2, ws.max_row + 1)):
            ability_en = ws.cell(row=row, column=name_en_col).value
            if ability_en:
                description = self.get_ability_description(ability_en)
                ws.cell(row=row, column=desc_col, value=description)

        # 保存更新后的文件
        print("保存更新后的Excel文件...")
        wb.save(excel_file)

def main():
    scraper = AbilityDescriptionScraper()
    scraper.update_excel_with_descriptions('output/ability_data.xlsx')

if __name__ == "__main__":
    main() 